from openpyxl import Workbook, load_workbook
dzienniklatprodukcji={"a":2010,"b":2011,"c":2012,"d":2013,"e":2014,"f":2015,}
licznik1=0
licznik2=0
licznik3=0
licznik4=0
licznik5=0
#zrob wybor normalnego pliku a zamiane w pdf
#moze jakies gui
#zapytaj sie czy kazdy linde ma rok produkcji w numerze 7 miejscu
#-------------------------------------------------------------------------------------------------
while 1:
    print("wybierz typ resursu: ")
    print("wozek, zuraw, podest, dzwignik, protokol")
    wybor=input()
    while 1:
        if wybor=="wozek" or "zuraw" or "podest" or "dzwignik" or "protokol":
            pass
        else:
            print("Wpisz nazwe poprawnie")
            break
        if wybor=="wozek":
            wb=load_workbook(filename="Resurswzor.xlsx")
            ws=wb.active
            licznik1+=1
            nazwapliku=f'Resurs{licznik1}.xlsx'
            ekspl=input("Eksploatujący: ")
            nrewi=input("Numer ewidencyjny: ")
            produ=input("Producent: ").lower()
            typ=input("Typ: ")
            nrser=input("Numer seryjny: ").lower()
            if produ=="linde":
                rok=dzienniklatprodukcji.get(nrser[6])
            else:
                rok=input("Rok produkcji: ")
            if produ=="linde":
                udzwig=str(typ[1])+str(typ[2])+"00"
            else:
                udzwig=input("Udźwig: ")
            mtg=input("Liczba mtg: ")
            wartoscgran=input("Wartość graniczna (domyślnie 40000): ")
            if wartoscgran=="":
                wartoscgran=40000
            wspolczynnik=str(input("Współczynnik (domyślny 1.0): "))
            wspolczynnik=wspolczynnik.replace(",",".")
            if wspolczynnik=="":
                wspolczynnik=1.0
            else:
                wspolczynnik=float(wspolczynnik)
            wartoscredu=str(input("Wartość redukująca 100 lub 125 procent: "))
            wartoscredu.replace("%","")
            wartoscredu=float(wartoscredu)
            if wartoscredu==100:
                wartoscredu=1
            else:
                wartoscredu=1.25
            razem=typ.upper()+" "+produ.capitalize()
            ws["d7"]=ekspl.capitalize()
            ws["d9"]=nrewi.upper()
            ws["l8"]=razem
            ws["l9"]=nrser.upper()
            ws["l10"]=rok
            ws["l11"]=udzwig
            ws["l12"]=mtg
            ws["i21"]=wspolczynnik
            ws["c21"]=wartoscgran
            ws["l21"]=wartoscredu
            wb.save(filename=nazwapliku)
            print("Jeżeli chcesz zrobić kolejny resurs wcisnij enter")
            koniec=input("Jeżeli chcesz zmienic typ resursu wpisz 'zmien'")
            if koniec=="zmien":
                break

        if wybor=="zuraw":
            wb=load_workbook(filename="Resurszurawiawzor.xlsx")
            ws=wb.active
            licznik2+=1
            nazwapliku=f'Resurszurawia{licznik2}.xlsx'
            ekspl=input("Eksploatujący: ")
            nrewi=input("Numer ewidencyjny: ")
            produ=input("Producent: ").lower()
            typ=input("Typ: ")
            nrser=input("Numer seryjny: ").lower()
            rok=input("Rok produkcji: ")
            udzwig=input("Udźwig: ")
            cykle=input("Ilość cykli: ")
            dnipracy=input("Ilość dni pracy (domyślnie 240): ")
            if dnipracy=="":
                dnipracy=240
            wartoscgran=input("Wartość graniczna (domyślnie 60000): ")
            if wartoscgran=="":
                wartoscgran=60000
            razem=typ.upper()+" "+produ.capitalize()
            ws["d7"]=ekspl.capitalize()
            ws["d9"]=nrewi.upper()
            ws["n8"]=razem
            ws["n9"]=nrser.upper()
            ws["n10"]=rok
            ws["n11"]=udzwig
            ws["n12"]=cykle
            ws["c22"]=dnipracy
            ws["g22"]=wartoscgran
            wb.save(filename=nazwapliku)
            print("Jeżeli chcesz zrobić kolejny resurs wcisnij enter")
            koniec=input("Jeżeli chcesz zmienic typ resursu wpisz 'zmien'")
            if koniec=="zmien":
                break

        if wybor=="podest":
            wb=load_workbook(filename="Resurspodestuwzor.xlsx")
            wb.iso_dates=True
            ws=wb.active
            licznik3+=1
            nazwapliku=f'Resurspodestu{licznik3}.xlsx'
            ekspl=input("Eksploatujący: ")
            nrewi=input("Numer ewidencyjny: ")
            produ=input("Producent: ").lower()
            typ=input("Typ: ")
            nrser=input("Numer seryjny: ").lower()
            rok=input("Rok produkcji: ")
            udzwig=input("Udźwig: ")
            cykle=input("Ilość cykli: ")
            dnipracy=input("Ilość dni pracy (domyślnie 240): ")
            if dnipracy=="":
                dnipracy=240
            wartoscgran=input("Wartość graniczna (domyślnie 60000): ")
            if wartoscgran=="":
                wartoscgran=60000
            razem=typ.upper()+" "+produ.capitalize()
            ws["d7"]=ekspl.capitalize()
            ws["d9"]=nrewi.upper()
            ws["n8"]=razem
            ws["n9"]=nrser.upper()
            ws["n10"]=rok
            ws["n11"]=udzwig
            ws["n12"]=cykle
            ws["c22"]=dnipracy
            ws["g22"]=wartoscgran
            wb.save(filename=nazwapliku)
            print("Jeżeli chcesz zrobić kolejny resurs wcisnij enter")
            koniec=input("Jeżeli chcesz zmienic typ resursu wpisz 'zmien'")
            if koniec=="zmien":
                break

        if wybor=="dzwignik":
            wb=load_workbook(filename="Resursdzwignikawzor.xlsx")
            wb.iso_dates=True
            ws=wb.active
            licznik4+=1
            nazwapliku=f'Resursdzwignika{licznik4}.xlsx'
            ekspl=input("Eksploatujący: ")
            nrewi=input("Numer ewidencyjny: ")
            produ=input("Producent: ").lower()
            typ=input("Typ: ")
            nrser=input("Numer seryjny: ").lower()
            rok=input("Rok produkcji: ")
            udzwig=input("Udźwig: ")
            cykle=input("Ilość cykli: ")
            dnipracy=input("Ilość dni pracy (domyślnie 240): ")
            if dnipracy=="":
                dnipracy=240
            wartoscgran=input("Wartość graniczna (domyślnie 60000): ")
            if wartoscgran=="":
                wartoscgran=60000
            razem=typ.upper()+" "+produ.capitalize()
            ws["d7"]=ekspl.capitalize()
            ws["d9"]=nrewi.upper()
            ws["n8"]=razem
            ws["n9"]=nrser.upper()
            ws["n10"]=rok
            ws["n11"]=udzwig
            ws["n12"]=cykle
            ws["c22"]=dnipracy
            ws["g22"]=wartoscgran
            wb.save(filename=nazwapliku)
            print("Jeżeli chcesz zrobić kolejny resurs wcisnij enter")
            koniec=input("Jeżeli chcesz zmienic typ resursu wpisz 'zmien'")
            if koniec=="zmien":
                break

        if wybor=="protokol":
            wb=load_workbook(filename="Protokolwzor.xlsx")
            wb.iso_dates=True
            ws=wb.active
            licznik5+=1
            nazwapliku=f'Protokol{licznik5}.xlsx'
            numer=input("Numer protokołu: ")
            nazwafirmy=input("Nazwa firmy: ")
            typ=input("Typ: ")
            nrfabr=input("Numer fabryczny: ").lower()
            stanl=input("Stan licznika: ")
            opis=input("Opis wykonywanych czynności: ")
            #czesci=input("")
            uwagi=input("Uwagi i zalecenia: ")
            wyko=input("Wykonał: ")
            czas=input("Czas pracy: ")
            dojazd=input("Dojazd: ")
            ws["c1"]=numer
            ws["b4"]=nazwafirmy.capitalize()
            ws["g4"]=typ.upper()
            ws["g6"]=nrfabr.upper()
            ws["g8"]=stanl
            ws["a12"]=opis
            #tu beda czesci
            ws["a36"]=uwagi
            ws["a42"]=wyko
            ws["d42"]=czas
            ws["d45"]=dojazd
            wb.save(filename=nazwapliku)
            print("Jeżeli chcesz zrobić kolejny protokół wcisnij enter")
            koniec=input("Jeżeli chcesz zmienic typ resursu wpisz 'zmien'")
            if koniec=="zmien":
                break