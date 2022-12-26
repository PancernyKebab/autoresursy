import os
from openpyxl import Workbook, load_workbook
import os
from win32com import client
import time
sciezka=os.getcwd()
sciezkawynik=os.getcwd()
sciezkawynik+="/wyniki"
dzienniklatprodukcji={"l":2000,"m":2001,"n":2002,"p":2003,"r":2004,"s":2005,"t":2006,"u":2007,"w":2008,"z":2009,"a":2010,"b":2011,"c":2012,"d":2013,"e":2014,"f":2015,"g":2016,"h":2017,"j":2018,"v":2019,"x":2020,"y":2021,"k":2022}
licznik1=0
licznik2=0
licznik3=0
licznik4=0
licznik5=0
p=11
a=os.getcwd()
#zrob zamiane na apk moze
#zrob wybor normalnego pliku a zamiane w pdf
#moze jakies gui
# jest zamiana na pdf w wózku ale nie dziala usuniecie xlsx nie wiadomo dlaczego
#-------------------------------------------------------------------------------------------------
cos=True
while cos:
    f=input("Dokumenty mają mieć format pdf czy xlsx").lower()
    if f=="pdf" or f=="xlsx":
        cos=False
    else:
        print("wpisz poprawnie")

if f=="pdf":
    while 1:
        print("wybierz typ dokumentu: ")
        print("wozek, zuraw, podest, dzwignik, protokol")
        wybor=input().lower()
        while 1:
            if wybor=="wozek":
                wb=load_workbook(filename="Resurswzor.xlsx")
                ws=wb.active
                licznik1+=1
                nazwapliku=f'Resurs{licznik1}.xlsx'
                nazwadopdf=f'Resurs{licznik1}'
                ekspl=input("Eksploatujący: ")
                nrewi=input("Numer ewidencyjny: ")
                produ=input("Producent: ").lower()
                typ=input("Typ: ")
                nrser=input("Numer seryjny: ").lower()
                if produ=="linde":
                    rok=dzienniklatprodukcji.get(nrser[6])
                    if rok=="NoneType":
                        rok=input("Rok produkcji: ")
                    else:
                        print(f'Czy to jest poprawny rok produkcji? {rok} (wciśnij enter jeżeli jest poprawny)')
                        tak=input().lower()
                        if tak=="":
                            pass
                        else:
                            rok=input("Rok produkcji: ")
                else:
                    rok=input("Rok produkcji: ")
                if produ=="linde":
                    udzwig=str(typ[1])+str(typ[2])+"00"
                else:
                    udzwig=input("Udźwig: ")
                mtg=input("Liczba mtg: ")
                wartoscgran=input("Wartość graniczna (domyślnie 40000): ")
                if wartoscgran=="":
                    wartoscgran=40000
                wspolczynnik=str(input("Współczynnik (domyślny 1.0): "))
                wspolczynnik=wspolczynnik.replace(",",".")
                try:
                    if wspolczynnik=="":
                        wspolczynnik=1.0
                    else:
                        wspolczynnik=float(wspolczynnik)
                except ValueError:
                    print("Została wpisana zła wartość współczynnika")
                wartoscredu=str(input("Wartość redukująca (domyślnie 100%)"))
                try:
                    if wartoscredu=="":
                        wartoscredu.replace("%","")
                        wartoscredu=1
                    else:
                        wartoscredu.replace("%","")
                        wartoscredu=f'{wartoscredu[0]}.{wartoscredu[-2]}{wartoscredu[-1]}'
                except IndexError:
                    print("Została wpisana zła wartość redukująca")
                razem=typ.upper()+" "+produ.capitalize()
                ws["d7"]=ekspl.capitalize()
                ws["d9"]=nrewi.upper()
                ws["l8"]=razem
                ws["l9"]=nrser.upper()
                ws["l10"]=rok
                ws["l11"]=udzwig
                ws["l12"]=mtg
                ws["i21"]=wspolczynnik
                ws["c21"]=wartoscgran
                ws["l21"]=wartoscredu
                wb.save(filename=nazwapliku)
                excel = client.Dispatch("Excel.Application")
                excel.Visible = False
                excel.DisplayAlerts = False
                sheets = excel.Workbooks.Open(f'{sciezka}\{nazwadopdf}')
                work_sheets = sheets.Worksheets[0]
                work_sheets.ExportAsFixedFormat(0, f'{sciezkawynik}\{nazwadopdf}')
                excel.quit()
                time.sleep(0.01)
                try:
                    os.remove(f'{sciezka}\{nazwapliku}')
                except PermissionError:
                    print("cos sie zepsulo")
                print("Jeżeli chcesz zrobić kolejny resurs wcisnij enter")
                koniec=input("Jeżeli chcesz zmienic typ resursu wpisz 'zmien'")
                if koniec=="zmien":
                    break

            elif wybor=="zuraw":
                wb=load_workbook(filename="Resurszurawiawzor.xlsx")
                ws=wb.active
                licznik2+=1
                nazwapliku=f'Resurszurawia{licznik2}.xlsx'
                ekspl=input("Eksploatujący: ")
                nrewi=input("Numer ewidencyjny: ")
                produ=input("Producent: ").lower()
                typ=input("Typ: ")
                nrser=input("Numer seryjny: ").lower()
                rok=input("Rok produkcji: ")
                udzwig=input("Udźwig: ")
                cykle=input("Ilość cykli: ")
                dnipracy=input("Ilość dni pracy (domyślnie 240): ")
                if dnipracy=="":
                    dnipracy=240
                wartoscgran=input("Wartość graniczna (domyślnie 60000): ")
                if wartoscgran=="":
                    wartoscgran=60000
                razem=typ.upper()+" "+produ.capitalize()
                ws["d7"]=ekspl.capitalize()
                ws["d9"]=nrewi.upper()
                ws["n8"]=razem
                ws["n9"]=nrser.upper()
                ws["n10"]=rok
                ws["n11"]=udzwig
                ws["n12"]=cykle
                ws["c22"]=dnipracy
                ws["g22"]=wartoscgran
                wb.save(filename=nazwapliku)
                print("Jeżeli chcesz zrobić kolejny resurs wcisnij enter")
                koniec=input("Jeżeli chcesz zmienic typ resursu wpisz 'zmien'")
                if koniec=="zmien":
                    break

            elif wybor=="podest":
                wb=load_workbook(filename="Resurspodestuwzor.xlsx")
                ws=wb.active
                licznik3+=1
                nazwapliku=f'Resurspodestu{licznik3}.xlsx'
                ekspl=input("Eksploatujący: ")
                nrewi=input("Numer ewidencyjny: ")
                produ=input("Producent: ").lower()
                typ=input("Typ: ")
                nrser=input("Numer seryjny: ").lower()
                rok=input("Rok produkcji: ")
                udzwig=input("Udźwig: ")
                cykle=input("Ilość cykli: ")
                dnipracy=input("Ilość dni pracy (domyślnie 240): ")
                if dnipracy=="":
                    dnipracy=240
                wartoscgran=input("Wartość graniczna (domyślnie 60000): ")
                if wartoscgran=="":
                    wartoscgran=60000
                razem=typ.upper()+" "+produ.capitalize()
                ws["d7"]=ekspl.capitalize()
                ws["d9"]=nrewi.upper()
                ws["n8"]=razem
                ws["n9"]=nrser.upper()
                ws["n10"]=rok
                ws["n11"]=udzwig
                ws["n12"]=cykle
                ws["c22"]=dnipracy
                ws["g22"]=wartoscgran
                wb.save(filename=nazwapliku)
                print("Jeżeli chcesz zrobić kolejny resurs wcisnij enter")
                koniec=input("Jeżeli chcesz zmienic typ resursu wpisz 'zmien'")
                if koniec=="zmien":
                    break

            elif wybor=="dzwignik":
                wb=load_workbook(filename="Resursdzwignikawzor.xlsx")
                ws=wb.active
                licznik4+=1
                nazwapliku=f'Resursdzwignika{licznik4}.xlsx'
                ekspl=input("Eksploatujący: ")
                nrewi=input("Numer ewidencyjny: ")
                produ=input("Producent: ").lower()
                typ=input("Typ: ")
                nrser=input("Numer seryjny: ").lower()
                rok=input("Rok produkcji: ")
                udzwig=input("Udźwig: ")
                cykle=input("Ilość cykli: ")
                dnipracy=input("Ilość dni pracy (domyślnie 240): ")
                if dnipracy=="":
                    dnipracy=240
                wartoscgran=input("Wartość graniczna (domyślnie 60000): ")
                if wartoscgran=="":
                    wartoscgran=60000
                razem=typ.upper()+" "+produ.capitalize()
                ws["d7"]=ekspl.capitalize()
                ws["d9"]=nrewi.upper()
                ws["n8"]=razem
                ws["n9"]=nrser.upper()
                ws["n10"]=rok
                ws["n11"]=udzwig
                ws["n12"]=cykle
                ws["c22"]=dnipracy
                ws["g22"]=wartoscgran
                wb.save(filename=nazwapliku)
                print("Jeżeli chcesz zrobić kolejny resurs wcisnij enter")
                koniec=input("Jeżeli chcesz zmienic typ resursu wpisz 'zmien'")
                if koniec=="zmien":
                    break

            elif wybor=="protokol":
                wb=load_workbook(filename="Protokolwzor.xlsx")
                ws=wb.active
                licznik5+=1
                nazwapliku=f'Protokol{licznik5}.xlsx'
                numer=input("Numer protokołu: ")
                nazwafirmy=input("Nazwa firmy: ")
                typ=input("Typ: ")
                nrfabr=input("Numer fabryczny: ").lower()
                stanl=input("Stan licznika: ")
                opis=input("Opis wykonywanych czynności: ")
                print("Wpisz nazwy i ilość części użytych, pozostaw puste jeżeli chcesz przejść dalej")
                while 1:
                    p+=1
                    czesc=input()
                    if czesc=="" or p==35:
                        break
                    ws[f'f{p}']=czesc
                uwagi=input("Uwagi i zalecenia: ")
                wyko=input("Wykonał: ")
                czas=input("Czas pracy: ")
                dojazd=input("Dojazd: ")
                ws["c1"]=numer
                ws["b4"]=nazwafirmy.capitalize()
                ws["g4"]=typ.upper()
                ws["g6"]=nrfabr.upper()
                ws["g8"]=stanl
                ws["a12"]=opis
                ws["a36"]=uwagi
                ws["a42"]=wyko
                ws["d42"]=czas
                ws["d45"]=dojazd
                wb.save(filename=nazwapliku)
                print("Jeżeli chcesz zrobić kolejny protokół wcisnij enter")
                koniec=input("Jeżeli chcesz zmienic typ resursu wpisz 'zmien'")
                if koniec=="zmien":
                    break
            else:
                print("wpisz poprawnie")
                break
else:
    while 1:
        print("wybierz typ dokumentu: ")
        print("wozek, zuraw, podest, dzwignik, protokol")
        wybor=input().lower()

        while 1:
            if wybor=="wozek":
                wb=load_workbook(filename="Resurswzor.xlsx")
                ws=wb.active
                licznik1+=1
                nazwapliku=f'Resurs{licznik1}.xlsx'
                ekspl=input("Eksploatujący: ")
                nrewi=input("Numer ewidencyjny: ")
                produ=input("Producent: ").lower()
                typ=input("Typ: ")
                nrser=input("Numer seryjny: ").lower()
                if produ=="linde":
                    rok=dzienniklatprodukcji.get(nrser[6])
                    if rok=="NoneType":
                        rok=input("Rok produkcji: ")
                    else:
                        print(f'Czy to jest poprawny rok produkcji? {rok} (wciśnij enter jeżeli jest poprawny)')
                        tak=input().lower()
                        if tak=="":
                            pass
                        else:
                            rok=input("Rok produkcji: ")
                else:
                    rok=input("Rok produkcji: ")
                if produ=="linde":
                    udzwig=str(typ[1])+str(typ[2])+"00"
                else:
                    udzwig=input("Udźwig: ")
                mtg=input("Liczba mtg: ")
                wartoscgran=input("Wartość graniczna (domyślnie 40000): ")
                if wartoscgran=="":
                    wartoscgran=40000
                wspolczynnik=str(input("Współczynnik (domyślny 1.0): "))
                wspolczynnik=wspolczynnik.replace(",",".")
                if wspolczynnik=="":
                    wspolczynnik=1.0
                else:
                    wspolczynnik=float(wspolczynnik)
                wartoscredu=str(input("Wartość redukująca (domyślnie 100%)"))
                if wartoscredu=="":
                    wartoscredu.replace("%","")
                    wartoscredu=1
                else:
                    wartoscredu.replace("%","")
                    wartoscredu=f'{wartoscredu[0]}.{wartoscredu[-2]}{wartoscredu[-1]}'
                razem=typ.upper()+" "+produ.capitalize()
                ws["d7"]=ekspl.capitalize()
                ws["d9"]=nrewi.upper()
                ws["l8"]=razem
                ws["l9"]=nrser.upper()
                ws["l10"]=rok
                ws["l11"]=udzwig
                ws["l12"]=mtg
                ws["i21"]=wspolczynnik
                ws["c21"]=wartoscgran
                ws["l21"]=wartoscredu
                wb.save(filename=nazwapliku)
                print("Jeżeli chcesz zrobić kolejny resurs wcisnij enter")
                koniec=input("Jeżeli chcesz zmienic typ resursu wpisz 'zmien'")
                if koniec=="zmien":
                    break

            elif wybor=="zuraw":
                wb=load_workbook(filename="Resurszurawiawzor.xlsx")
                ws=wb.active
                licznik2+=1
                nazwapliku=f'Resurszurawia{licznik2}.xlsx'
                ekspl=input("Eksploatujący: ")
                nrewi=input("Numer ewidencyjny: ")
                produ=input("Producent: ").lower()
                typ=input("Typ: ")
                nrser=input("Numer seryjny: ").lower()
                rok=input("Rok produkcji: ")
                udzwig=input("Udźwig: ")
                cykle=input("Ilość cykli: ")
                dnipracy=input("Ilość dni pracy (domyślnie 240): ")
                if dnipracy=="":
                    dnipracy=240
                wartoscgran=input("Wartość graniczna (domyślnie 60000): ")
                if wartoscgran=="":
                    wartoscgran=60000
                razem=typ.upper()+" "+produ.capitalize()
                ws["d7"]=ekspl.capitalize()
                ws["d9"]=nrewi.upper()
                ws["n8"]=razem
                ws["n9"]=nrser.upper()
                ws["n10"]=rok
                ws["n11"]=udzwig
                ws["n12"]=cykle
                ws["c22"]=dnipracy
                ws["g22"]=wartoscgran
                wb.save(filename=nazwapliku)
                print("Jeżeli chcesz zrobić kolejny resurs wcisnij enter")
                koniec=input("Jeżeli chcesz zmienic typ resursu wpisz 'zmien'")
                if koniec=="zmien":
                    break

            elif wybor=="podest":
                wb=load_workbook(filename="Resurspodestuwzor.xlsx")
                ws=wb.active
                licznik3+=1
                nazwapliku=f'Resurspodestu{licznik3}.xlsx'
                ekspl=input("Eksploatujący: ")
                nrewi=input("Numer ewidencyjny: ")
                produ=input("Producent: ").lower()
                typ=input("Typ: ")
                nrser=input("Numer seryjny: ").lower()
                rok=input("Rok produkcji: ")
                udzwig=input("Udźwig: ")
                cykle=input("Ilość cykli: ")
                dnipracy=input("Ilość dni pracy (domyślnie 240): ")
                if dnipracy=="":
                    dnipracy=240
                wartoscgran=input("Wartość graniczna (domyślnie 60000): ")
                if wartoscgran=="":
                    wartoscgran=60000
                razem=typ.upper()+" "+produ.capitalize()
                ws["d7"]=ekspl.capitalize()
                ws["d9"]=nrewi.upper()
                ws["n8"]=razem
                ws["n9"]=nrser.upper()
                ws["n10"]=rok
                ws["n11"]=udzwig
                ws["n12"]=cykle
                ws["c22"]=dnipracy
                ws["g22"]=wartoscgran
                wb.save(filename=nazwapliku)
                print("Jeżeli chcesz zrobić kolejny resurs wcisnij enter")
                koniec=input("Jeżeli chcesz zmienic typ resursu wpisz 'zmien'")
                if koniec=="zmien":
                    break

            elif wybor=="dzwignik":
                wb=load_workbook(filename="Resursdzwignikawzor.xlsx")
                ws=wb.active
                licznik4+=1
                nazwapliku=f'Resursdzwignika{licznik4}.xlsx'
                ekspl=input("Eksploatujący: ")
                nrewi=input("Numer ewidencyjny: ")
                produ=input("Producent: ").lower()
                typ=input("Typ: ")
                nrser=input("Numer seryjny: ").lower()
                rok=input("Rok produkcji: ")
                udzwig=input("Udźwig: ")
                cykle=input("Ilość cykli: ")
                dnipracy=input("Ilość dni pracy (domyślnie 240): ")
                if dnipracy=="":
                    dnipracy=240
                wartoscgran=input("Wartość graniczna (domyślnie 60000): ")
                if wartoscgran=="":
                    wartoscgran=60000
                razem=typ.upper()+" "+produ.capitalize()
                ws["d7"]=ekspl.capitalize()
                ws["d9"]=nrewi.upper()
                ws["n8"]=razem
                ws["n9"]=nrser.upper()
                ws["n10"]=rok
                ws["n11"]=udzwig
                ws["n12"]=cykle
                ws["c22"]=dnipracy
                ws["g22"]=wartoscgran
                wb.save(filename=nazwapliku)
                print("Jeżeli chcesz zrobić kolejny resurs wcisnij enter")
                koniec=input("Jeżeli chcesz zmienic typ resursu wpisz 'zmien'")
                if koniec=="zmien":
                    break

            elif wybor=="protokol":
                wb=load_workbook(filename="Protokolwzor.xlsx")
                ws=wb.active
                licznik5+=1
                nazwapliku=f'Protokol{licznik5}.xlsx'
                numer=input("Numer protokołu: ")
                nazwafirmy=input("Nazwa firmy: ")
                typ=input("Typ: ")
                nrfabr=input("Numer fabryczny: ").lower()
                stanl=input("Stan licznika: ")
                opis=input("Opis wykonywanych czynności: ")
                print("Wpisz nazwy i ilość części użytych, pozostaw puste jeżeli chcesz przejść dalej")
                while 1:
                    p+=1
                    czesc=input()
                    if czesc=="" or p==35:
                        break
                    ws[f'f{p}']=czesc
                uwagi=input("Uwagi i zalecenia: ")
                wyko=input("Wykonał: ")
                czas=input("Czas pracy: ")
                dojazd=input("Dojazd: ")
                ws["c1"]=numer
                ws["b4"]=nazwafirmy.capitalize()
                ws["g4"]=typ.upper()
                ws["g6"]=nrfabr.upper()
                ws["g8"]=stanl
                ws["a12"]=opis
                ws["a36"]=uwagi
                ws["a42"]=wyko
                ws["d42"]=czas
                ws["d45"]=dojazd
                wb.save(filename=nazwapliku)
                print("Jeżeli chcesz zrobić kolejny protokół wcisnij enter")
                koniec=input("Jeżeli chcesz zmienic typ resursu wpisz 'zmien'")
                if koniec=="zmien":
                    break
            else:
                print("wpisz poprawnie")
                break