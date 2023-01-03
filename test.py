from openpyxl import workbook, load_workbook
import win32com
import os
desktop = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop') 
a=os.getcwd()
a=a+"\wyniki"
wb=load_workbook("Resurswzor.xlsx")
ws=wb.active
ws["l8"]="cos"
print(a, desktop)
f=r"C:\Users\Dell\Desktop"
#wb.save("j.xlsx",f)
wb.save(os.path.join(f+"\j.xlsx"))
