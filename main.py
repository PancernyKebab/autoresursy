import os
from openpyxl import Workbook, load_workbook
from win32com import client
import time
from datetime import date, datetime
#potrzeba jeszcze zainstalować Pillow żeby obrazy sie zapisały
#----------------------------------------------------
excel = client.Dispatch("Excel.Application")
excel.quit()
testowy="plikxlsxdousuniecia.xlsx"
testowybezrozsz="plikxlsxdousuniecia"
sciezka=os.getcwd()
sciezkanapulpit = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop') 
dzienniklatprodukcji={"l":2000,"m":2001,"n":2002,"p":2003,"r":2004,"s":2005,"t":2006,"u":2007,"w":2008,"z":2009,"a":2010,"b":2011,"c":2012,"d":2013,"e":2014,"f":2015,"g":2016,"h":2017,"j":2018,"v":2019,"x":2020,"y":2021,"k":2022}
sc=os.path.join(sciezkanapulpit+"\wynikiprogramu")

p=11
czyprzerwac=0
dzis=date.today()
teraz=datetime.now().strftime("%H-%M-%S")
def usuniecie():
    try:
        #usuwa plik xlsx ktory byl potrzebny do stworzenia pdfa
        os.remove(os.path.join(f'{sc}\{testowy}'))
    except PermissionError:
        print("cos sie zepsulo")
def zmianatypu():
    while 1:
        print("Jeżeli chcesz zrobić kolejny resurs wcisnij enter")
        print("Jeżeli chcesz zmienic typ resursu wpisz 'zmien'")
        koniec=input()
        if koniec=="zmien":
            czyprzerwac=1
            return czyprzerwac
        elif koniec=="":
            czyprzerwac=2
            return czyprzerwac
        else:
            czyprzerwac=0
            pass
    return czyprzerwac
def dopdfa():
    wb.save(os.path.join(f'{sc}\{testowy}'))
    excel = client.Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    sheets = excel.Workbooks.Open(f'{sc}\{testowybezrozsz}')
    work_sheets = sheets.Worksheets[0]
    work_sheets.ExportAsFixedFormat(0, f'{sc}\{nazwadopdf}')
    excel.quit()
#-------------------------------------------------------------------------------------------------

cos=True
#tworze folder z wynikami programu na pulpicie
try:
    os.mkdir(sc)
except FileExistsError:
    pass
#wybor miedzy pdfem a xlsx
while cos:
    f=input("Dokumenty mają mieć format pdf czy xlsx: ").lower()
    if f=="pdf" or f=="xlsx":
        cos=False
    else:
        print("wpisz poprawnie")

try:
    if f=="pdf":
        while 1:
            print("wybierz typ dokumentu: ")
            print("wozek, zuraw, podest, dzwignik, protokol")
            wybor=input().lower()
            ekspl=wybor
            while 1:
                if wybor=="wozek":
                    #laduje wzor z pliku xlsx
                    wb=load_workbook(os.path.join(sciezka+"\Resurswzor.xlsx"))
                    ws=wb.active
                    ekspl=input("Eksploatujący: ")
                    if ekspl=="":
                        ekspl=wybor
                    nrewi=input("Numer ewidencyjny: ")
                    produ=input("Producent: ").lower()
                    typ=input("Typ: ")
                    nrser=input("Numer seryjny: ").lower()
                    if produ=="linde":
                        rok=dzienniklatprodukcji.get(nrser[6])
                        if rok=="NoneType":
                            rok=input("Rok produkcji: ")
                        else:
                            print(f'Czy to jest poprawny rok produkcji? {rok} (wciśnij enter jeżeli jest poprawny)')
                            tak=input().lower()
                            if tak=="" or tak=="tak":
                                pass
                            else:
                                rok=input("Rok produkcji: ")
                    else:
                        rok=input("Rok produkcji: ")
                    if produ=="linde":
                        udzwig=str(typ[1])+str(typ[2])+"00"
                    else:
                        udzwig=input("Udźwig: ")
                    mtg=input("Liczba mtg: ")
                    wartoscgran=input("Wartość graniczna (domyślnie 40000): ")
                    if wartoscgran=="":
                        wartoscgran=40000
                    wspolczynnik=str(input("Współczynnik (domyślny 1.0): "))
                    wspolczynnik=wspolczynnik.replace(",",".")
                    try:
                        if wspolczynnik=="":
                            wspolczynnik=1.0
                        else:
                            wspolczynnik=float(wspolczynnik)
                    except ValueError:
                        print("Została wpisana zła wartość współczynnika")
                    wartoscredu=str(input("Wartość redukująca (domyślnie 100%)"))
                    try:
                        if wartoscredu=="":
                            wartoscredu.replace("%","")
                            wartoscredu=1
                        else:
                            wartoscredu.replace("%","")
                            wartoscredu=f'{wartoscredu[0]}.{wartoscredu[-2]}{wartoscredu[-1]}'
                    except IndexError:
                        print("Została wpisana zła wartość redukująca")
                    razem=typ.upper()+" "+produ.capitalize()
                    #wpisuje dane do podanych komorek w pliku xlsx
                    ws["d7"]=ekspl.capitalize()
                    ws["d9"]=nrewi.upper()
                    ws["l8"]=razem
                    ws["l9"]=nrser.upper()
                    ws["l10"]=rok
                    ws["l11"]=udzwig
                    ws["l12"]=mtg
                    ws["i21"]=wspolczynnik
                    ws["c21"]=wartoscgran
                    ws["l21"]=wartoscredu
                    nazwapliku=f'{ekspl}-{dzis}-{teraz}.xlsx'
                    nazwadopdf=f'{ekspl}-{dzis}-{teraz}'
                    print(ekspl,dzis,teraz)
                    #tworze plik xlsx ktory posluzy do stworzenia pdfa
                    dopdfa()
                    time.sleep(0.2)
                    usuniecie()
                    zmiana1=zmianatypu()
                    if zmiana1==1:
                        break
                    elif zmiana1==2:
                        pass

                elif wybor=="zuraw":
                    wb=load_workbook(os.path.join(sciezka+"\Resurszurawiawzor.xlsx"))
                    ws=wb.active
                    nazwapliku=f'{ekspl}-{dzis}-{teraz}.xlsx'
                    nazwadopdf=f'{ekspl}-{dzis}-{teraz}'
                    ekspl=input("Eksploatujący: ")
                    nrewi=input("Numer ewidencyjny: ")
                    produ=input("Producent: ").lower()
                    typ=input("Typ: ")
                    nrser=input("Numer seryjny: ").lower()
                    rok=input("Rok produkcji: ")
                    udzwig=input("Udźwig: ")
                    cykle=input("Ilość cykli: ")
                    dnipracy=input("Ilość dni pracy (domyślnie 240): ")
                    if dnipracy=="":
                        dnipracy=240
                    wartoscgran=input("Wartość graniczna (domyślnie 60000): ")
                    if wartoscgran=="":
                        wartoscgran=60000
                    razem=typ.upper()+" "+produ.capitalize()
                    ws["d7"]=ekspl.capitalize()
                    ws["d9"]=nrewi.upper()
                    ws["n8"]=razem
                    ws["n9"]=nrser.upper()
                    ws["n10"]=rok
                    ws["n11"]=udzwig
                    ws["n12"]=cykle
                    ws["c22"]=dnipracy
                    ws["g22"]=wartoscgran
                    dopdfa()
                    time.sleep(0.2)
                    usuniecie()
                    zmiana1=zmianatypu()
                    if zmiana1==1:
                        break
                    elif zmiana1==2:
                        pass

                elif wybor=="podest":
                    wb=load_workbook(os.path.join(sciezka+"\Resurspodestuwzor.xlsx"))
                    ws=wb.active
                    nazwapliku=f'{ekspl}-{dzis}-{teraz}.xlsx'
                    nazwadopdf=f'{ekspl}-{dzis}-{teraz}'
                    ekspl=input("Eksploatujący: ")
                    nrewi=input("Numer ewidencyjny: ")
                    produ=input("Producent: ").lower()
                    typ=input("Typ: ")
                    nrser=input("Numer seryjny: ").lower()
                    rok=input("Rok produkcji: ")
                    udzwig=input("Udźwig: ")
                    cykle=input("Ilość cykli: ")
                    dnipracy=input("Ilość dni pracy (domyślnie 240): ")
                    if dnipracy=="":
                        dnipracy=240
                    wartoscgran=input("Wartość graniczna (domyślnie 60000): ")
                    if wartoscgran=="":
                        wartoscgran=60000
                    razem=typ.upper()+" "+produ.capitalize()
                    ws["d7"]=ekspl.capitalize()
                    ws["d9"]=nrewi.upper()
                    ws["n8"]=razem
                    ws["n9"]=nrser.upper()
                    ws["n10"]=rok
                    ws["n11"]=udzwig
                    ws["n12"]=cykle
                    ws["c22"]=dnipracy
                    ws["g22"]=wartoscgran
                    dopdfa()
                    time.sleep(0.2)
                    usuniecie()
                    zmiana1=zmianatypu()
                    if zmiana1==1:
                        break
                    elif zmiana1==2:
                        pass

                elif wybor=="dzwignik":
                    wb=load_workbook(os.path.join(sciezka+"\Resursdzwignikawzor.xlsx"))
                    ws=wb.active
                    nazwapliku=f'{ekspl}-{dzis}-{teraz}.xlsx'
                    nazwadopdf=f'{ekspl}-{dzis}-{teraz}'
                    ekspl=input("Eksploatujący: ")
                    nrewi=input("Numer ewidencyjny: ")
                    produ=input("Producent: ").lower()
                    typ=input("Typ: ")
                    nrser=input("Numer seryjny: ").lower()
                    rok=input("Rok produkcji: ")
                    udzwig=input("Udźwig: ")
                    cykle=input("Ilość cykli: ")
                    dnipracy=input("Ilość dni pracy (domyślnie 240): ")
                    if dnipracy=="":
                        dnipracy=240
                    wartoscgran=input("Wartość graniczna (domyślnie 60000): ")
                    if wartoscgran=="":
                        wartoscgran=60000
                    razem=typ.upper()+" "+produ.capitalize()
                    ws["d7"]=ekspl.capitalize()
                    ws["d9"]=nrewi.upper()
                    ws["n8"]=razem
                    ws["n9"]=nrser.upper()
                    ws["n10"]=rok
                    ws["n11"]=udzwig
                    ws["n12"]=cykle
                    ws["c22"]=dnipracy
                    ws["g22"]=wartoscgran
                    dopdfa()
                    time.sleep(0.2)
                    usuniecie()
                    zmiana1=zmianatypu()
                    if zmiana1==1:
                        break
                    elif zmiana1==2:
                        pass

                elif wybor=="protokol":
                    wb=load_workbook(os.path.join(sciezka+"\Protokolwzor.xlsx"))
                    ws=wb.active
                    nazwapliku=f'{ekspl}-{dzis}-{teraz}.xlsx'
                    nazwadopdf=f'{ekspl}-{dzis}-{teraz}'
                    numer=input("Numer protokołu: ")
                    nazwafirmy=input("Nazwa firmy: ")
                    typ=input("Typ: ")
                    nrfabr=input("Numer fabryczny: ").lower()
                    stanl=input("Stan licznika: ")
                    opis=input("Opis wykonywanych czynności: ")
                    print("Wpisz nazwy i ilość części użytych, pozostaw puste jeżeli chcesz przejść dalej")
                    while 1:
                        p+=1
                        czesc=input()
                        if czesc=="" or p==35:
                            break
                        ws[f'f{p}']=czesc
                    uwagi=input("Uwagi i zalecenia: ")
                    wyko=input("Wykonał: ")
                    czas=input("Początek pracy (np. 15:00): ")
                    czas2=input("koniec pracy: ")
                    dojazd=input("Dojazd: ")
                    ws["c1"]=numer
                    ws["b4"]=nazwafirmy.capitalize()
                    ws["g4"]=typ.upper()
                    ws["g6"]=nrfabr.upper()
                    ws["g8"]=stanl
                    ws["a12"]=opis
                    ws["a36"]=uwagi
                    ws["a42"]=wyko
                    ws["d43"]=czas
                    ws["f43"]=czas2
                    ws["d45"]=dojazd
                    dopdfa()
                    time.sleep(0.2)
                    usuniecie()
                    zmiana1=zmianatypu()
                    if zmiana1==1:
                        break
                    elif zmiana1==2:
                        pass

                else:
                    print("wpisz poprawnie")
                    break
    else:
        while 1:

            print("wybierz typ dokumentu: ")
            print("wozek, zuraw, podest, dzwignik, protokol")
            wybor=input().lower()

            while 1:
                if wybor=="wozek":
                    wb=load_workbook(filename="Resurswzor.xlsx")
                    ws=wb.active
                    nazwapliku=f'{ekspl}-{dzis}-{teraz}.xlsx'
                    ekspl=input("Eksploatujący: ")
                    nrewi=input("Numer ewidencyjny: ")
                    produ=input("Producent: ").lower()
                    typ=input("Typ: ")
                    nrser=input("Numer seryjny: ").lower()
                    if produ=="linde":
                        rok=dzienniklatprodukcji.get(nrser[6])
                        if rok=="NoneType":
                            rok=input("Rok produkcji: ")
                        else:
                            print(f'Czy to jest poprawny rok produkcji? {rok} (wciśnij enter jeżeli jest poprawny)')
                            tak=input().lower()
                            if tak=="":
                                pass
                            else:
                                rok=input("Rok produkcji: ")
                    else:
                        rok=input("Rok produkcji: ")
                    if produ=="linde":
                        udzwig=str(typ[1])+str(typ[2])+"00"
                    else:
                        udzwig=input("Udźwig: ")
                    mtg=input("Liczba mtg: ")
                    wartoscgran=input("Wartość graniczna (domyślnie 40000): ")
                    if wartoscgran=="":
                        wartoscgran=40000
                    wspolczynnik=str(input("Współczynnik (domyślny 1.0): "))
                    wspolczynnik=wspolczynnik.replace(",",".")
                    try:
                        if wspolczynnik=="":
                            wspolczynnik=1.0
                        else:
                            wspolczynnik=float(wspolczynnik)
                    except ValueError:
                        print("została wpisana zla wartosc współczynnika")
                    wartoscredu=str(input("Wartość redukująca (domyślnie 100%)"))
                    try:
                        if wartoscredu=="":
                            wartoscredu.replace("%","")
                            wartoscredu=1
                        else:
                            wartoscredu.replace("%","")
                            wartoscredu=f'{wartoscredu[0]}.{wartoscredu[-2]}{wartoscredu[-1]}'
                    except IndexError:
                        print("Została wpisana zła wartość redukująca")
                    razem=typ.upper()+" "+produ.capitalize()
                    ws["d7"]=ekspl.capitalize()
                    ws["d9"]=nrewi.upper()
                    ws["l8"]=razem
                    ws["l9"]=nrser.upper()
                    ws["l10"]=rok
                    ws["l11"]=udzwig
                    ws["l12"]=mtg
                    ws["i21"]=wspolczynnik
                    ws["c21"]=wartoscgran
                    ws["l21"]=wartoscredu
                    wb.save(os.path.join(sc+"\\"+nazwapliku))
                    zmiana1=zmianatypu()
                    if zmiana1==1:
                        break
                    elif zmiana1==2:
                        pass

                elif wybor=="zuraw":
                    wb=load_workbook(filename="Resurszurawiawzor.xlsx")
                    ws=wb.active
                    nazwapliku=f'{ekspl}-{dzis}-{teraz}.xlsx'
                    ekspl=input("Eksploatujący: ")
                    nrewi=input("Numer ewidencyjny: ")
                    produ=input("Producent: ").lower()
                    typ=input("Typ: ")
                    nrser=input("Numer seryjny: ").lower()
                    rok=input("Rok produkcji: ")
                    udzwig=input("Udźwig: ")
                    cykle=input("Ilość cykli: ")
                    dnipracy=input("Ilość dni pracy (domyślnie 240): ")
                    if dnipracy=="":
                        dnipracy=240
                    wartoscgran=input("Wartość graniczna (domyślnie 60000): ")
                    if wartoscgran=="":
                        wartoscgran=60000
                    razem=typ.upper()+" "+produ.capitalize()
                    ws["d7"]=ekspl.capitalize()
                    ws["d9"]=nrewi.upper()
                    ws["n8"]=razem
                    ws["n9"]=nrser.upper()
                    ws["n10"]=rok
                    ws["n11"]=udzwig
                    ws["n12"]=cykle
                    ws["c22"]=dnipracy
                    ws["g22"]=wartoscgran
                    wb.save(os.path.join(sc+"\\"+nazwapliku))
                    zmiana1=zmianatypu()
                    if zmiana1==1:
                        break
                    elif zmiana1==2:
                        pass


                elif wybor=="podest":
                    wb=load_workbook(filename="Resurspodestuwzor.xlsx")
                    ws=wb.active
                    nazwapliku=f'{ekspl}-{dzis}-{teraz}.xlsx'
                    ekspl=input("Eksploatujący: ")
                    nrewi=input("Numer ewidencyjny: ")
                    produ=input("Producent: ").lower()
                    typ=input("Typ: ")
                    nrser=input("Numer seryjny: ").lower()
                    rok=input("Rok produkcji: ")
                    udzwig=input("Udźwig: ")
                    cykle=input("Ilość cykli: ")
                    dnipracy=input("Ilość dni pracy (domyślnie 240): ")
                    if dnipracy=="":
                        dnipracy=240
                    wartoscgran=input("Wartość graniczna (domyślnie 60000): ")
                    if wartoscgran=="":
                        wartoscgran=60000
                    razem=typ.upper()+" "+produ.capitalize()
                    ws["d7"]=ekspl.capitalize()
                    ws["d9"]=nrewi.upper()
                    ws["n8"]=razem
                    ws["n9"]=nrser.upper()
                    ws["n10"]=rok
                    ws["n11"]=udzwig
                    ws["n12"]=cykle
                    ws["c22"]=dnipracy
                    ws["g22"]=wartoscgran
                    wb.save(os.path.join(sc+"\\"+nazwapliku))
                    zmiana1=zmianatypu()
                    if zmiana1==1:
                        break
                    elif zmiana1==2:
                        pass


                elif wybor=="dzwignik":
                    wb=load_workbook(filename="Resursdzwignikawzor.xlsx")
                    ws=wb.active
                    nazwapliku=f'{ekspl}-{dzis}-{teraz}.xlsx'
                    ekspl=input("Eksploatujący: ")
                    nrewi=input("Numer ewidencyjny: ")
                    produ=input("Producent: ").lower()
                    typ=input("Typ: ")
                    nrser=input("Numer seryjny: ").lower()
                    rok=input("Rok produkcji: ")
                    udzwig=input("Udźwig: ")
                    cykle=input("Ilość cykli: ")
                    dnipracy=input("Ilość dni pracy (domyślnie 240): ")
                    if dnipracy=="":
                        dnipracy=240
                    wartoscgran=input("Wartość graniczna (domyślnie 60000): ")
                    if wartoscgran=="":
                        wartoscgran=60000
                    razem=typ.upper()+" "+produ.capitalize()
                    ws["d7"]=ekspl.capitalize()
                    ws["d9"]=nrewi.upper()
                    ws["n8"]=razem
                    ws["n9"]=nrser.upper()
                    ws["n10"]=rok
                    ws["n11"]=udzwig
                    ws["n12"]=cykle
                    ws["c22"]=dnipracy
                    ws["g22"]=wartoscgran
                    wb.save(os.path.join(sc+"\\"+nazwapliku))
                    zmiana1=zmianatypu()
                    if zmiana1==1:
                        break
                    elif zmiana1==2:
                        pass


                elif wybor=="protokol":
                    wb=load_workbook(filename="Protokolwzor.xlsx")
                    ws=wb.active
                    nazwapliku=f'{ekspl}-{dzis}-{teraz}.xlsx'
                    numer=input("Numer protokołu: ")
                    nazwafirmy=input("Nazwa firmy: ")
                    typ=input("Typ: ")
                    nrfabr=input("Numer fabryczny: ").lower()
                    stanl=input("Stan licznika: ")
                    opis=input("Opis wykonywanych czynności: ")
                    print("Wpisz nazwy i ilość części użytych, pozostaw puste jeżeli chcesz przejść dalej")
                    while 1:
                        p+=1
                        czesc=input()
                        if czesc=="" or p==35:
                            break
                        ws[f'f{p}']=czesc
                    uwagi=input("Uwagi i zalecenia: ")
                    wyko=input("Wykonał: ")
                    czas=input("Początek pracy (np. 15:00): ")
                    czas2=input("koniec pracy: ")
                    dojazd=input("Dojazd: ")
                    ws["c1"]=numer
                    ws["b4"]=nazwafirmy.capitalize()
                    ws["g4"]=typ.upper()
                    ws["g6"]=nrfabr.upper()
                    ws["g8"]=stanl
                    ws["a12"]=opis
                    ws["a36"]=uwagi
                    ws["a42"]=wyko
                    ws["d43"]=czas
                    ws["f43"]=czas2
                    ws["d45"]=dojazd
                    wb.save(os.path.join(sc+"\\"+nazwapliku))
                    zmiana1=zmianatypu()
                    if zmiana1==1:
                        break
                    elif zmiana1==2:
                        pass

                else:
                    print("wpisz poprawnie")
                    break


except Exception as e:
    try:
        with open(f'{sc}\\bledy.txt',"x")as plik:
            plik.write(f'błąd: {str(e)}\n')
            print("Znaleziono błąd program sie zamyka D:")
    except:
        with open(f'{sc}\\bledy.txt',"a")as plik:
            plik.write(f'błąd: {str(e)}\n')
            print("Znaleziono błąd program sie zamyka D:")